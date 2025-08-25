#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Nmap Results to Excel Converter

This script processes Nmap scan results from various formats (.gnmap, .nmap, .xml)
and generates a professional Excel report matching the original script's output format.

Features:
- Supports multiple Nmap output formats
- Uses actual service names from nmap output
- Professional Excel formatting
- Error handling and logging
- Popup notifications

Author: Security Team
Version: 2.2
"""

import os
import sys
import logging
import pandas as pd
import re
import xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tkinter as tk
from tkinter import messagebox

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Fallback port services mapping for cases where service name is not detected
PORT_SERVICES = {
    1: "TCPMUX", 5: "RJE", 7: "ECHO", 9: "DISCARD", 11: "SYSTAT", 13: "DAYTIME", 17: "QOTD", 18: "MSP",
    19: "CHARGEN", 20: "FTP-DATA", 21: "FTP", 22: "SSH", 23: "TELNET", 25: "SMTP", 37: "TIME",
    42: "NAMESERVER", 43: "NICNAME", 49: "TACACS", 50: "RE-MAIL-CK", 53: "DOMAIN",
    67: "BOOTPS", 68: "BOOTPC", 69: "TFTP", 70: "GOPHER", 71: "NETRJS-1", 72: "NETRJS-2",
    73: "NETRJS-3", 74: "NETRJS-4", 79: "FINGER", 80: "HTTP", 88: "KERBEROS", 95: "SUPDUP",
    101: "HOSTNAME", 102: "ISO-TSAP", 105: "CSNET-NS", 107: "RTELNET", 109: "POP2", 110: "POP3",
    111: "SUNRPC", 113: "AUTH", 115: "SFTP", 117: "UUCP-PATH", 119: "NNTP", 123: "NTP",
    135: "MS-RPC", 137: "NETBIOS-NS", 138: "NETBIOS-DGM", 139: "NETBIOS-SSN", 143: "IMAP",
    161: "SNMP", 162: "SNMPTRAP", 163: "CMIP-MAN", 164: "CMIP-AGENT", 174: "MAILQ",
    177: "XDMCP", 178: "NEXTSTEP", 179: "BGP", 191: "PROSPERO", 194: "IRC",
    199: "SMUX", 201: "AT-RTMP", 202: "AT-NBP", 204: "AT-ECHO", 206: "AT-ZIS",
    209: "QMTP", 210: "Z39.50", 213: "IPX", 220: "IMAP3", 245: "LINK", 347: "FATSERV",
    363: "RSVP_TUNNEL", 369: "RPC2PORTMAP", 370: "CODAAUTH2", 372: "ULISTPROC", 389: "LDAP",
    427: "SLP", 434: "MOBILE-IP", 435: "MOBILIP-MN", 443: "HTTPS", 444: "SNPP",
    445: "MICROSOFT-DS", 464: "KPASSWD", 468: "PHOTURIS", 487: "SAFT", 488: "GSS-HTTP",
    496: "PIM-RP-DISC", 500: "ISAKMP", 512: "EXEC", 513: "LOGIN", 514: "SHELL",
    515: "PRINTER", 517: "TALK", 518: "NTALK", 520: "EFS", 521: "RIPNG", 525: "TIMED",
    526: "TEMPO", 530: "COURIER", 531: "CONFERENCE", 532: "NETNEWS", 533: "NETWALL",
    538: "GDOMAP", 540: "UUCP", 543: "KLOGIN", 544: "KSHELL", 546: "DHCPV6-CLIENT",
    547: "DHCPV6-SERVER", 548: "AFPOVERTCP", 549: "IDFP", 554: "RTSP", 556: "REMOTEFS",
    563: "NNTPS", 565: "WHOAMI", 587: "SUBMISSION", 593: "HTTP-RPC-EPMAP", 616: "SCOHELP",
    617: "SCO-INETMGR", 625: "APPLE-XSRVR-ADMIN", 631: "IPP", 636: "LDAPS",
    646: "LDP", 648: "RRP", 666: "DOOM", 667: "DISCLOSE", 668: "MECOMM", 683: "CORBA-IIOP",
    687: "ASIPREGISTRY", 691: "RESVC", 700: "ELCSD", 705: "AGENTX", 711: "CISCO-TDP",
    714: "IRIS-XPCS", 749: "KERBEROS-ADM", 750: "KERBEROS-IV", 751: "KERBEROS_MASTER",
    752: "PASSWD_SERVER", 754: "KREG", 760: "KRBUPDATE", 765: "WEBSTER", 767: "PHONEBOOK",
    808: "OMIRR", 873: "RSYNC", 901: "SAMBA-SWAT", 989: "FTPS-DATA", 990: "FTPS",
    992: "TELNETS", 993: "IMAPS", 994: "IRCS", 995: "POP3S", 1025: "NFS", 1080: "SOCKS",
    1194: "OPENVPN", 1241: "NESSUS", 1311: "RXE", 1433: "MS-SQL-S", 1434: "MS-SQL-M",
    1521: "ORACLE", 1604: "ICABROWSER", 1723: "PPTP", 1741: "CICSLISTENER",
    1777: "PHAROS", 1883: "MQTT", 1900: "UPNP", 1911: "STARMAN", 2000: "CISCO-SCCP",
    2049: "NFS", 2082: "CPANEL", 2083: "CPANEL-SSL", 2100: "AMIGANETFS", 2222: "ETC-SSH",
    2375: "DOCKER", 2376: "DOCKER-S", 2483: "ORACLE-SSL", 2484: "ORACLE-SSL",
    2601: "ZEBRA", 2604: "OSPFD", 2628: "DICT", 3000: "PPPD", 3128: "SQUID",
    3260: "ISCSI", 3306: "MYSQL", 3389: "MS-WBT-SERVER", 3456: "VAT", 3478: "STUN",
    3632: "DISTCC", 3689: "DAAP", 3690: "SVN", 4000: "REMOTEANYTHING", 4369: "EPMD",
    4899: "RADMIN", 5000: "UPnP", 5001: "COMMPLEX-MAIN", 5060: "SIP", 5190: "AOL",
    5357: "WSDAPI", 5432: "POSTGRESQL", 5631: "PC-ANYWHERE", 5666: "NRPE",
    5800: "VNC-HTTP", 5900: "VNC", 5901: "VNC-1", 5985: "WSMAN", 6000: "X11",
    6566: "SANE-PORT", 6588: "ANALOGX", 6665: "IRCU", 8000: "HTTP-ALT", 8008: "HTTP",
    8009: "AJPV13", 8080: "HTTP-PROXY", 8081: "BLACKICE-ICECAP", 8443: "HTTPS-ALT",
    8888: "NEWRELIC", 9100: "JETDIRECT", 9999: "ABYSS", 10000: "WEBMIN",
    10001: "SNET-SENSOR-MGMT", 32768: "FILENET-TMS", 49152: "UNKNOWN-49152"
}

def get_fallback_service_name(port):
    """Returns the fallback service name for a given port if nmap didn't detect it."""
    try:
        port = int(port)
        
        if port in PORT_SERVICES:
            return PORT_SERVICES[port]
        
        if 0 <= port <= 1023:
            return "RESERVED"
        elif 1024 <= port <= 49151:
            return "REGISTERED"
        elif 49152 <= port <= 65535:
            return "DYNAMIC/PRIVATE"
        else:
            return "INVALID"
            
    except (ValueError, TypeError):
        return "UNKNOWN"

def show_popup_message(title, message, message_type="info"):
    """Show popup message using tkinter."""
    try:
        # Create root window but hide it
        root = tk.Tk()
        root.withdraw()
        
        if message_type == "info":
            messagebox.showinfo(title, message)
        elif message_type == "warning":
            messagebox.showwarning(title, message)
        elif message_type == "error":
            messagebox.showerror(title, message)
        
        root.destroy()
    except Exception as e:
        # Fallback to console if GUI not available
        print(f"{title}: {message}")

def parse_gnmap_file(file_path):
    """Parse .gnmap (greppable) format and extract actual service names."""
    results = []
    
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            if line.startswith('Host:') and 'Ports:' in line:
                # Extract IP address
                ip_match = re.search(r'Host:\s+(\S+)', line)
                if not ip_match:
                    continue
                    
                ip = ip_match.group(1)
                
                # Extract ports section
                ports_match = re.search(r'Ports:\s+(.+?)(?:\s+Ignored|$)', line)
                if not ports_match:
                    continue
                    
                ports_section = ports_match.group(1)
                
                # Parse individual ports
                port_entries = ports_section.split(',')
                for entry in port_entries:
                    entry = entry.strip()
                    if not entry:
                        continue
                        
                    # Parse port entry: port/state/protocol/owner/service/SunRPC/version
                    parts = entry.split('/')
                    if len(parts) >= 5:
                        port = parts[0].strip()
                        state = parts[1].strip()
                        protocol = parts[2].strip()
                        service = parts[4].strip() if len(parts) > 4 else ''
                        
                        if state == 'open':
                            try:
                                port_num = int(port)
                                # Use actual service name from nmap, fallback if empty or unknown
                                if not service or service in ['', 'unknown', '?']:
                                    service = get_fallback_service_name(port_num)
                                
                                results.append({
                                    'Host': ip,
                                    'Port': port_num,
                                    'Protocol': protocol,
                                    'Service': service.upper(),
                                    'State': state
                                })
                            except ValueError:
                                continue
    
    return results

def parse_xml_file(file_path):
    """Parse .xml format and extract actual service names."""
    results = []
    
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        
        for host in root.findall('.//host'):
            # Get IP address
            address_elem = host.find('.//address[@addrtype="ipv4"]')
            if address_elem is None:
                continue
                
            ip = address_elem.get('addr')
            
            # Get open ports
            for port in host.findall('.//port'):
                state_elem = port.find('state')
                if state_elem is not None and state_elem.get('state') == 'open':
                    port_num = int(port.get('portid'))
                    protocol = port.get('protocol')
                    
                    # Extract service name from XML
                    service_elem = port.find('service')
                    if service_elem is not None:
                        service = service_elem.get('name', '')
                        # Also check for product info
                        product = service_elem.get('product', '')
                        if product and product != service:
                            service = f"{service} ({product})" if service else product
                    else:
                        service = ''
                    
                    # Use fallback if no service detected
                    if not service or service in ['', 'unknown', '?']:
                        service = get_fallback_service_name(port_num)
                    
                    results.append({
                        'Host': ip,
                        'Port': port_num,
                        'Protocol': protocol,
                        'Service': service.upper(),
                        'State': 'open'
                    })
                    
    except ET.ParseError as e:
        logger.error(f"Error parsing XML file {file_path}: {e}")
    except Exception as e:
        logger.error(f"Unexpected error parsing XML file {file_path}: {e}")
    
    return results

def parse_nmap_file(file_path):
    """Parse .nmap (normal) format and extract actual service names."""
    results = []
    current_ip = None
    
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
        
    # Find all host blocks
    host_blocks = re.findall(r'Nmap scan report for ([^\n]+).*?(?=Nmap scan report for|\Z)', content, re.DOTALL)
    
    for block in host_blocks:
        lines = block.split('\n')
        
        # Extract IP from first line
        ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', lines[0])
        if not ip_match:
            continue
            
        current_ip = ip_match.group(1)
        
        # Look for port information
        in_port_section = False
        for line in lines[1:]:
            line = line.strip()
            
            # Check if we're in the port/state/service section
            if 'PORT' in line and 'STATE' in line and 'SERVICE' in line:
                in_port_section = True
                continue
                
            if in_port_section and line:
                # Parse port line: "22/tcp   open  ssh     OpenSSH 7.4"
                port_match = re.match(r'(\d+)/(tcp|udp)\s+open\s+(\S+)(?:\s+(.+))?', line)
                if port_match:
                    port_num = int(port_match.group(1))
                    protocol = port_match.group(2)
                    service = port_match.group(3) if port_match.group(3) else ''
                    version_info = port_match.group(4) if port_match.group(4) else ''
                    
                    # Combine service with version info if available
                    if version_info and version_info.strip():
                        service = f"{service} ({version_info.strip()})"
                    
                    # Use fallback if no service detected
                    if not service or service in ['', 'unknown', '?']:
                        service = get_fallback_service_name(port_num)
                    
                    results.append({
                        'Host': current_ip,
                        'Port': port_num,
                        'Protocol': protocol,
                        'Service': service.upper(),
                        'State': 'open'
                    })
            elif in_port_section and not line:
                # Empty line might indicate end of port section
                in_port_section = False
    
    return results

def style_excel_worksheet(worksheet):
    """Apply professional styling to the Excel worksheet."""
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Style header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Style data rows
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # Auto-adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Set row height for better readability
    for row_idx in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 20

def process_nmap_files(input_directory, output_file):
    """Process all Nmap files in the directory and generate Excel report."""
    try:
        logger.info(f"Processing Nmap files from {input_directory}")
        
        all_results = []
        
        # Find all Nmap files
        nmap_files = []
        for file in os.listdir(input_directory):
            if file.endswith(('.gnmap', '.nmap', '.xml')):
                nmap_files.append(os.path.join(input_directory, file))
        
        if not nmap_files:
            logger.error("No Nmap files found in directory")
            show_popup_message(
                "No Files Found", 
                "No Nmap files (.gnmap, .nmap, .xml) found in the current directory.\n\nPlease ensure you have Nmap scan results in the same folder as this script.",
                "warning"
            )
            return False
        
        logger.info(f"Found {len(nmap_files)} Nmap files")
        
        # Process each file
        for file_path in nmap_files:
            logger.info(f"Processing {os.path.basename(file_path)}")
            
            if file_path.endswith('.gnmap'):
                results = parse_gnmap_file(file_path)
            elif file_path.endswith('.xml'):
                results = parse_xml_file(file_path)
            elif file_path.endswith('.nmap'):
                results = parse_nmap_file(file_path)
            else:
                continue
                
            all_results.extend(results)
            logger.info(f"Found {len(results)} open ports in {os.path.basename(file_path)}")
        
        if not all_results:
            logger.warning("No open ports found in any files")
            show_popup_message(
                "No Open Ports Found", 
                "No open ports were found in any of the processed Nmap files.\n\nThis could mean:\n• All ports are filtered/closed\n• The scan results are empty\n• The files may be corrupted or incomplete",
                "info"
            )
            return False
        
        # Convert to DataFrame
        df = pd.DataFrame(all_results)
        logger.info(f"Total open ports found: {len(df)}")
        
        # Create formatted port info using actual service names
        df['Formatted Port Info'] = df.apply(
            lambda row: f"{row['Protocol'].upper()}/{int(row['Port'])}/{row['Service']}/OPEN",
            axis=1
        )
        
        # Group by host
        merged_df = df.groupby('Host')['Formatted Port Info'].apply(
            lambda x: '\n'.join(sorted(x.unique()))
        ).reset_index()
        
        # Rename columns
        merged_df.rename(columns={
            'Host': 'Address of Host (Hostname)',
            'Formatted Port Info': 'Protocol/Port/Service/Status'
        }, inplace=True)
        
        # Add comments
        merged_df['Comments'] = (
            "All ports on the target host which are not listed here were observed to be in state "
            "as FILTERED. Port range scanned: 1-65534."
        )
        
        # Sort by host
        merged_df.sort_values(by='Address of Host (Hostname)', inplace=True)
        
        # Save to Excel
        merged_df.to_excel(output_file, index=False, engine='openpyxl')
        logger.info(f"Data saved to {output_file}")
        
        # Apply styling
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active
        style_excel_worksheet(worksheet)
        
        # Add metadata
        worksheet.title = "Nmap Port Scan Results"
        worksheet.sheet_properties.tabColor = "4472C4"
        
        workbook.save(output_file)
        logger.info(f"Excel styling completed")
        
        # Show success popup
        hosts_count = len(merged_df)
        ports_count = len(df)
        show_popup_message(
            "Success!", 
            f"Nmap results processed successfully!\n\n• Found {hosts_count} host(s) with open ports\n• Total open ports: {ports_count}\n• Results saved to: NMap_Port_Scan_Result.xlsx",
            "info"
        )
        
        return True
        
    except Exception as e:
        logger.error(f"Error processing Nmap files: {str(e)}", exc_info=True)
        show_popup_message(
            "Error", 
            f"An error occurred while processing the files:\n\n{str(e)}\n\nCheck the log for more details.",
            "error"
        )
        return False

def main():
    """Main function."""
    try:
        logger.info("Nmap Results to Excel Converter starting...")
        
        # Get script directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Output file
        output_file = os.path.join(script_dir, 'NMap_Port_Scan_Result.xlsx')
        
        # Process files
        success = process_nmap_files(script_dir, output_file)
        
        if success:
            logger.info("Task completed successfully.")
            print("SUCCESS: Nmap results processed and saved to NMap_Port_Scan_Result.xlsx")
            return 0
        else:
            logger.error("Task failed.")
            print("ERROR: Failed to process Nmap results. Check log for details.")
            return 1
            
    except Exception as e:
        logger.critical(f"Unhandled exception: {str(e)}", exc_info=True)
        print(f"CRITICAL ERROR: {str(e)}")
        show_popup_message(
            "Critical Error", 
            f"A critical error occurred:\n\n{str(e)}\n\nThe application will now exit.",
            "error"
        )
        return 1

if __name__ == "__main__":
    sys.exit(main())